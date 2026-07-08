from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

from services.excel_exporter import export_excel
from services.excel_parser import extract_segments
from services.translation_align import is_source_already_target_language


def _make_workbook() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Install Guide"
    sheet["A1"] = "Module"
    sheet["A2"] = "12345"

    hidden = workbook.create_sheet("Hidden Notes")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "Keep me"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_extract_segments_includes_hidden_sheet_titles_and_cells() -> None:
    content = _make_workbook()

    segments = extract_segments(content)

    assert [segment["segment_type"] for segment in segments] == [
        "sheet_title",
        "cell",
        "sheet_title",
        "cell",
    ]
    assert segments[0]["sheet"] == "Install Guide"
    assert segments[1]["cell"] == "A1"
    assert segments[2]["sheet"] == "Hidden Notes"
    assert segments[3]["cell"] == "A1"


def test_export_excel_translates_sheet_titles_and_cells() -> None:
    original = _make_workbook()
    segments = extract_segments(original)

    translated = export_excel(
        original,
        segments,
        [
            {"translation": "Installation Guide"},
            {"translation": "Module Item"},
            {"translation": "Hidden Notes EN"},
            {"translation": "Preserve me"},
        ],
    )

    workbook = openpyxl.load_workbook(BytesIO(translated))
    assert workbook.sheetnames == ["Installation Guide", "Hidden Notes EN"]
    assert workbook["Installation Guide"]["A1"].value == "Module Item"
    assert workbook["Hidden Notes EN"]["A1"].value == "Preserve me"
    assert workbook["Hidden Notes EN"].sheet_state == "hidden"


def test_mixed_language_segment_is_not_treated_as_already_translated() -> None:
    assert is_source_already_target_language("Level", "en-US") is True
    assert is_source_already_target_language("\u5185\u5bb9 Description", "en-US") is False


def test_export_excel_syncs_table_headers_with_translated_cells() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "\u5185\u5bb9"
    sheet["B1"] = "\u8fdb\u5ea6"
    sheet["A2"] = "\u793a\u4f8b"
    sheet["B2"] = "1"
    table = Table(displayName="Table1", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    output = BytesIO()
    workbook.save(output)
    original = output.getvalue()
    segments = extract_segments(original)

    translated = export_excel(
        original,
        segments,
        [
            {"translation": "Sheet1"},
            {"translation": "Content"},
            {"translation": "Progress"},
            {"translation": "Example"},
        ],
    )

    reloaded = openpyxl.load_workbook(BytesIO(translated))
    ws = reloaded["Sheet1"]
    reloaded_table = ws.tables["Table1"]
    assert ws["A1"].value == "Content"
    assert ws["B1"].value == "Progress"
    assert [column.name for column in reloaded_table.tableColumns] == ["Content", "Progress"]
